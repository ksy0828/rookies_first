import schedule
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import smtplib, os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()


def crawl_and_sendmail():
    url = "https://www.malware-traffic-analysis.net/2024/index.html"
    headers = {'User-Agent': 'Mozilla/5.0', 'Content-Type': 'text/html; charset=utf-8'}
    req = requests.get(url, headers=headers)
    soup = BeautifulSoup(req.text, "lxml")

    tags = soup.select('#main_content > div.blog_entry > ul > li > a.main_menu')

    wb = Workbook()
    ws = wb.active
    ws['A1'] = "제목"
    ws['B1'] = "링크"

    i = 2
    for tag in tags:
        ws.cell(row=i, column=1, value=tag.text)
        ws.cell(row=i, column=2, value=f"https://www.malware-traffic-analysis.net/2023/{tag['href']}")
        i += 1

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"malwares_{today}.xlsx"
    wb.save(filename)

    send_email(filename)

def send_email(file_path):
    send_email = os.getenv("SEND_MAIL") 
    send_pwd = os.getenv("SEND_PW") 
    recv_email = "kim42348@gmail.com"

    smtp = smtplib.SMTP('smtp.naver.com', 587)
    smtp.ehlo()
    smtp.starttls()
    smtp.login(send_email, send_pwd)

    text = "오늘의 최신 보안 동향 정보입니다. 첨부파일을 확인하세요."
    
    msg = MIMEMultipart()
    msg['Subject'] = f"malware-traffic-Report_{datetime.now().strftime('%Y-%m-%d')}"
    msg['From'] = send_email
    msg['To'] = recv_email

    contentPart = MIMEText(text)
    msg.attach(contentPart)

    # 파일 첨부
    with open(file_path, 'rb') as f:
        etc_part = MIMEApplication(f.read())
        etc_part.add_header('Content-Disposition', f'attachment; filename={file_path}')
        msg.attach(etc_part)

    # 이메일 전송
    email_string = msg.as_string()
    smtp.sendmail(send_email, recv_email, email_string)
    smtp.quit()


schedule.every().day.at("09:00").do(crawl_and_sendmail)

while True:
    schedule.run_pending()
    time.sleep(60)  # 1분마다 확인

# crawl_and_sendmail()